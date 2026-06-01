from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import math
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
import xlrd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls", ".json"}
SKIP_DIRECTORIES = {
    "__pycache__",
    ".git",
    ".idea",
    ".vscode",
    "build",
    "dist",
    "example",
    "examples",
    "output",
}

RAW_COLUMNS = [
    "publish_date",
    "article_title",
    "source_file",
    "reads",
    "delivered",
    "completion_rate",
    "avg_stay_seconds",
    "shares",
    "likes",
    "in_views",
    "favorites",
    "comments",
    "follows",
    "reward_yuan",
    "goods_yuan",
    "first_share_people",
    "total_share_people",
    "share_generated_reads",
    "recommend_reads",
    "search_reads",
    "message_reads",
    "chat_reads",
    "home_reads",
    "moments_reads",
    "friend_watch_reads",
    "other_reads",
]

RAW_DISPLAY = {
    "publish_date": "发布日期",
    "article_title": "文章标题",
    "source_file": "源文件",
    "reads": "阅读人数",
    "delivered": "送达人数",
    "completion_rate": "完读率",
    "avg_stay_seconds": "平均停留时长(秒)",
    "shares": "分享人数",
    "likes": "点赞人数",
    "in_views": "在看人数",
    "favorites": "收藏人数",
    "comments": "评论条数",
    "follows": "关注人数",
    "reward_yuan": "赞赏金额(元)",
    "goods_yuan": "带货金额(元)",
    "first_share_people": "首次分享人数",
    "total_share_people": "总分享人数",
    "share_generated_reads": "分享带来阅读",
    "recommend_reads": "推荐阅读",
    "search_reads": "搜一搜阅读",
    "message_reads": "公众号消息阅读",
    "chat_reads": "聊天会话阅读",
    "home_reads": "公众号主页阅读",
    "moments_reads": "朋友圈阅读",
    "friend_watch_reads": "朋友在看阅读",
    "other_reads": "其他来源阅读",
}

ANALYSIS_HEADERS = [
    "发布日期",
    "月份",
    "文章标题",
    "阅读人数",
    "送达人数",
    "打开率",
    "完读率",
    "平均停留时长(秒)",
    "分享人数",
    "点赞人数",
    "在看人数",
    "收藏人数",
    "评论条数",
    "关注人数",
    "赞赏金额(元)",
    "带货金额(元)",
    "分享率",
    "点赞率",
    "在看率",
    "收藏率",
    "评论率",
    "关注率",
    "千读赞赏(元)",
    "千读带货(元)",
    "阅读表现分",
    "互动表现分",
    "留存表现分",
    "转化表现分",
    "综合指数",
    "综合等级",
    "综合排名",
]

HELPER_HEADERS = [
    "阅读人数分位",
    "打开率分位",
    "分享率分位",
    "点赞率分位",
    "在看率分位",
    "收藏率分位",
    "评论率分位",
    "完读率分位",
    "停留时长分位",
    "关注率分位",
    "综合指数分位",
]

DETAIL_FIELD_ALIASES = {
    "阅读(人)": "reads",
    "阅读(次)": "reads",
    "阅读人数": "reads",
    "平均停留时长(秒)": "avg_stay_seconds",
    "完读率": "completion_rate",
    "阅读后关注(人)": "follows",
    "阅读后关注（人）": "follows",
    "分享(人)": "shares",
    "分享(次)": "shares",
    "点赞(人)": "likes",
    "点赞(次)": "likes",
    "在看(人)": "in_views",
    "在看(次)": "in_views",
    "收藏(人)": "favorites",
    "收藏(次)": "favorites",
    "评论(条)": "comments",
    "评论（条）": "comments",
    "赞赏(分)": "reward_fen",
    "送达人数": "delivered",
    "公众号消息阅读人数": "message_conversion_reads",
    "公众号消息阅读次数": "message_conversion_reads",
    "首次分享人数": "first_share_people",
    "首次分享次数": "first_share_people",
    "总分享人数": "total_share_people",
    "总分享次数": "total_share_people",
    "分享产生的阅读人数": "share_generated_reads",
    "分享产生的阅读次数": "share_generated_reads",
}

GENERIC_COLUMN_ALIASES = {
    "发布日期": "publish_date",
    "日期": "publish_date",
    "文章标题": "article_title",
    "标题": "article_title",
    "源文件": "source_file",
    "阅读人数": "reads",
    "阅读(人)": "reads",
    "送达人数": "delivered",
    "完读率": "completion_rate",
    "平均停留时长(秒)": "avg_stay_seconds",
    "分享人数": "shares",
    "分享(人)": "shares",
    "点赞人数": "likes",
    "点赞(人)": "likes",
    "在看人数": "in_views",
    "在看(人)": "in_views",
    "收藏人数": "favorites",
    "收藏(人)": "favorites",
    "评论条数": "comments",
    "评论（条）": "comments",
    "评论(条)": "comments",
    "关注人数": "follows",
    "阅读后关注（人）": "follows",
    "赞赏金额(元)": "reward_yuan",
    "带货金额(元)": "goods_yuan",
    "首次分享人数": "first_share_people",
    "总分享人数": "total_share_people",
    "分享带来阅读": "share_generated_reads",
    "分享产生的阅读人数": "share_generated_reads",
    "推荐阅读": "recommend_reads",
    "搜一搜阅读": "search_reads",
    "公众号消息阅读": "message_reads",
    "聊天会话阅读": "chat_reads",
    "公众号主页阅读": "home_reads",
    "朋友圈阅读": "moments_reads",
    "朋友在看阅读": "friend_watch_reads",
    "其他来源阅读": "other_reads",
}

READ_SOURCE_ALIASES = {
    "全部": "all",
    "推荐": "recommend_reads",
    "搜一搜": "search_reads",
    "公众号消息": "message_reads",
    "公众号主页": "home_reads",
    "聊天会话": "chat_reads",
    "朋友圈": "moments_reads",
    "朋友在看": "friend_watch_reads",
    "其他": "other_reads",
}


@dataclass(frozen=True)
class AnalysisOutputs:
    excel_path: Path
    html_path: Path
    total_files: int
    imported_articles: int
    skipped_duplicates: int
    failed_files: int


def get_app_root() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path.cwd().resolve()


def get_desktop_root() -> Path:
    desktop = Path.home() / "Desktop"
    return desktop if desktop.exists() else Path.home()


def choose_default_folder(name: str) -> Path:
    candidates = [get_app_root() / name, get_desktop_root() / name]
    for candidate in candidates:
        try:
            candidate.mkdir(parents=True, exist_ok=True)
            return candidate
        except Exception:
            continue
    return candidates[-1]


def normalize_text(value: Any) -> str:
    text = str(value or "").replace("\u3000", " ").replace("\xa0", " ")
    text = text.replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_label(value: Any) -> str:
    return normalize_text(value).replace(" ", "")


def safe_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and pd.notna(value):
        return float(value)
    text = normalize_text(value)
    if not text:
        return None
    percent = text.endswith("%")
    cleaned = text.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
    if not match:
        return None
    number = float(match.group(0))
    return number / 100 if percent else number


def safe_int(value: Any) -> int | None:
    number = safe_float(value)
    if number is None:
        return None
    return int(round(number))


def coalesce(*values: Any) -> Any:
    for value in values:
        if value is None:
            continue
        if isinstance(value, float) and math.isnan(value):
            continue
        if value == "":
            continue
        return value
    return None


def safe_div(numerator: Any, denominator: Any) -> float:
    num = safe_float(numerator)
    den = safe_float(denominator)
    if num is None or den in (None, 0):
        return 0.0
    return num / den


def percentile_series(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.notna().sum() <= 1:
        return pd.Series([50.0] * len(series), index=series.index)
    return numeric.rank(method="average", pct=True) * 100.0


def score_to_grade(value: float) -> str:
    if value >= 90:
        return "S"
    if value >= 75:
        return "A"
    if value >= 55:
        return "B"
    if value >= 35:
        return "C"
    return "D"


def parse_date(value: Any) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = normalize_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_date_from_path(path: Path) -> dt.date | None:
    year = next((int(part) for part in reversed(path.parts) if re.fullmatch(r"20\d{2}", part)), None)
    month_day = re.match(r"(\d{2})(\d{2})", path.stem)
    if year and month_day:
        return dt.date(year, int(month_day.group(1)), int(month_day.group(2)))
    return None


def sheet_cell_text(sheet: xlrd.sheet.Sheet, row: int, col: int) -> str:
    try:
        return normalize_text(sheet.cell_value(row, col))
    except Exception:
        return ""


def is_wechat_detail_sheet(sheet: xlrd.sheet.Sheet) -> bool:
    labels = {"数据概况", "阅读转化", "数据趋势明细", "阅读(人)", "完读率"}
    hits = 0
    for row in range(min(sheet.nrows, 40)):
        for col in range(min(sheet.ncols, 6)):
            if sheet_cell_text(sheet, row, col) in labels:
                hits += 1
                if hits >= 2:
                    return True
    return False


def map_detail_label(label: str) -> str | None:
    normalized = normalize_label(label)
    for alias, field in DETAIL_FIELD_ALIASES.items():
        if normalize_label(alias) == normalized:
            return field
    return None


def map_read_source(source_name: str) -> str | None:
    normalized = normalize_label(source_name)
    for alias, field in READ_SOURCE_ALIASES.items():
        if normalize_label(alias) == normalized:
            return field
    return None


def find_value_to_right(sheet: xlrd.sheet.Sheet, row: int, col: int) -> float | None:
    for current in range(col + 1, min(sheet.ncols, col + 4)):
        value = sheet.cell_value(row, current)
        parsed = safe_float(value)
        if parsed is not None:
            return parsed
    return None


def find_goods_amount(sheet: xlrd.sheet.Sheet) -> float | None:
    for row in range(min(sheet.nrows, 120)):
        for col in range(min(sheet.ncols, 6)):
            if "带货成交" not in normalize_text(sheet.cell_value(row, col)):
                continue
            for inner_row in range(row + 1, min(sheet.nrows, row + 12)):
                for inner_col in range(min(sheet.ncols, 6)):
                    if normalize_label(sheet.cell_value(inner_row, inner_col)) == "金额":
                        return find_value_to_right(sheet, inner_row, inner_col)
            return None
    return None


def find_trend_header(sheet: xlrd.sheet.Sheet) -> tuple[int, int, int, int, int | None] | None:
    for row in range(sheet.nrows):
        labels = [normalize_label(sheet.cell_value(row, col)) for col in range(min(sheet.ncols, 8))]
        if "日期" not in labels and "时间" not in labels:
            continue
        if "传播渠道" not in labels and "阅读来源" not in labels and "渠道" not in labels:
            continue
        if "阅读人数" not in labels and "阅读(人)" not in labels:
            continue
        date_col = next(i for i, label in enumerate(labels) if label in {"日期", "时间"})
        source_col = next(i for i, label in enumerate(labels) if label in {"传播渠道", "阅读来源", "渠道"})
        reads_col = next(i for i, label in enumerate(labels) if label in {"阅读人数", "阅读(人)"})
        shares_col = next((i for i, label in enumerate(labels) if label in {"分享人数", "分享(人)"}), None)
        return row, date_col, source_col, reads_col if shares_col is None else reads_col, shares_col
    return None


def parse_trend_rows(sheet: xlrd.sheet.Sheet) -> list[dict[str, Any]]:
    header = find_trend_header(sheet)
    if header is None:
        return []
    if len(header) == 5:
        header_row, date_col, source_col, reads_col, shares_col = header
    else:
        header_row, date_col, source_col, reads_col = header
        shares_col = None

    rows: list[dict[str, Any]] = []
    empty_streak = 0
    for row in range(header_row + 1, sheet.nrows):
        date_value = parse_date(sheet.cell_value(row, date_col))
        source = sheet_cell_text(sheet, row, source_col)
        reads = safe_float(sheet.cell_value(row, reads_col))
        shares = safe_float(sheet.cell_value(row, shares_col)) if shares_col is not None else None

        if not source and date_value is None and reads is None:
            empty_streak += 1
            if empty_streak >= 3:
                break
            continue
        empty_streak = 0
        if source and date_value:
            rows.append(
                {
                    "date": date_value,
                    "source": source,
                    "reads": reads or 0.0,
                    "shares": shares or 0.0,
                }
            )
    return rows


def parse_wechat_detail_xls(path: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    book = xlrd.open_workbook(str(path), on_demand=True)
    sheet = book.sheet_by_index(0)

    record = {column: None for column in RAW_COLUMNS}
    record["source_file"] = path.name
    record["article_title"] = normalize_text(sheet.cell_value(0, 1)) or path.stem

    for row in range(min(sheet.nrows, 120)):
        for col in range(min(sheet.ncols, 6)):
            field = map_detail_label(sheet_cell_text(sheet, row, col))
            if not field or field in record and record[field] is not None:
                continue
            record[field] = find_value_to_right(sheet, row, col)

    reward_fen = safe_float(record.pop("reward_fen", None)) if "reward_fen" in record else None
    record["reward_yuan"] = reward_fen / 100 if reward_fen is not None else 0.0
    record["goods_yuan"] = coalesce(find_goods_amount(sheet), 0.0)

    trend_rows = parse_trend_rows(sheet)
    record["publish_date"] = coalesce(
        min((row["date"] for row in trend_rows if map_read_source(row["source"]) == "all"), default=None),
        min((row["date"] for row in trend_rows), default=None),
        parse_date_from_path(path),
    )

    source_totals = {field: 0.0 for field in READ_SOURCE_ALIASES.values() if field != "all"}
    for row in trend_rows:
        field = map_read_source(row["source"])
        if field and field != "all":
            source_totals[field] += safe_float(row["reads"]) or 0.0
    record.update(source_totals)

    book.release_resources()
    return record, trend_rows


def normalize_generic_frame(frame: pd.DataFrame, source_name: str) -> pd.DataFrame:
    renamed = {}
    for column in frame.columns:
        if column in GENERIC_COLUMN_ALIASES:
            renamed[column] = GENERIC_COLUMN_ALIASES[column]
    normalized = frame.rename(columns=renamed).copy()
    if "source_file" not in normalized.columns:
        normalized["source_file"] = source_name
    if "article_title" not in normalized.columns:
        normalized["article_title"] = source_name

    for column in RAW_COLUMNS:
        if column not in normalized.columns:
            normalized[column] = None

    normalized["reward_yuan"] = normalized["reward_yuan"].fillna(
        pd.to_numeric(normalized.get("reward_fen"), errors="coerce").fillna(0) / 100
        if "reward_fen" in normalized.columns
        else 0
    )
    normalized["goods_yuan"] = pd.to_numeric(normalized["goods_yuan"], errors="coerce").fillna(0)
    normalized["publish_date"] = pd.to_datetime(normalized["publish_date"], errors="coerce").dt.date

    return normalized[RAW_COLUMNS]


def read_generic_tabular_file(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        frame = None
        for encoding in ("utf-8-sig", "gbk", "utf-8"):
            try:
                frame = pd.read_csv(path, encoding=encoding)
                break
            except Exception:
                continue
        if frame is None:
            frame = pd.read_csv(path)
    elif path.suffix.lower() == ".xlsx":
        frame = pd.read_excel(path)
    else:
        with path.open("r", encoding="utf-8") as file:
            payload = json.load(file)
        if isinstance(payload, list):
            frame = pd.DataFrame(payload)
        elif isinstance(payload, dict) and isinstance(payload.get("data"), list):
            frame = pd.DataFrame(payload["data"])
        elif isinstance(payload, dict):
            frame = pd.DataFrame([payload])
        else:
            raise ValueError(f"不支持的 JSON 结构: {type(payload)}")

    normalized = normalize_generic_frame(frame, path.name)
    return normalized.to_dict("records")


def read_file_records(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if path.suffix.lower() == ".xls":
        book = xlrd.open_workbook(str(path), on_demand=True)
        sheet = book.sheet_by_index(0)
        detail = is_wechat_detail_sheet(sheet)
        book.release_resources()
        if detail:
            record, trend_rows = parse_wechat_detail_xls(path)
            return [record], trend_rows
        frame = pd.read_excel(path)
        return normalize_generic_frame(frame, path.name).to_dict("records"), []
    return read_generic_tabular_file(path), []


def scan_input_files(input_dir: Path) -> list[Path]:
    files: list[Path] = []
    for path in sorted(input_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        relative_parts = path.relative_to(input_dir).parts[:-1]
        lowered_parts = {part.lower() for part in relative_parts}
        if lowered_parts & SKIP_DIRECTORIES:
            continue
        files.append(path)
    return files


def build_signature(record: dict[str, Any]) -> str:
    payload = {
        "publish_date": record.get("publish_date").isoformat() if isinstance(record.get("publish_date"), dt.date) else record.get("publish_date"),
        "article_title": normalize_text(record.get("article_title")),
        "reads": safe_int(record.get("reads")),
        "delivered": safe_int(record.get("delivered")),
        "completion_rate": round(safe_float(record.get("completion_rate")) or 0.0, 6),
        "avg_stay_seconds": round(safe_float(record.get("avg_stay_seconds")) or 0.0, 3),
        "shares": safe_int(record.get("shares")),
        "likes": safe_int(record.get("likes")),
        "in_views": safe_int(record.get("in_views")),
        "favorites": safe_int(record.get("favorites")),
        "comments": safe_int(record.get("comments")),
        "follows": safe_int(record.get("follows")),
        "total_share_people": safe_int(record.get("total_share_people")),
        "share_generated_reads": safe_int(record.get("share_generated_reads")),
    }
    content = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    return hashlib.sha1(content.encode("utf-8")).hexdigest()


def coerce_raw_frame(frame: pd.DataFrame) -> pd.DataFrame:
    raw = frame.copy()
    for column in RAW_COLUMNS:
        if column not in raw.columns:
            raw[column] = None

    numeric_columns = [column for column in RAW_COLUMNS if column not in {"publish_date", "article_title", "source_file"}]
    for column in numeric_columns:
        raw[column] = pd.to_numeric(raw[column], errors="coerce").fillna(0)

    raw["article_title"] = raw["article_title"].map(normalize_text)
    raw["source_file"] = raw["source_file"].map(normalize_text)
    raw["publish_date"] = pd.to_datetime(raw["publish_date"], errors="coerce").dt.date
    raw["reward_yuan"] = raw["reward_yuan"].fillna(0)
    raw["goods_yuan"] = raw["goods_yuan"].fillna(0)
    return raw[RAW_COLUMNS]


def calculate_metrics(raw_df: pd.DataFrame) -> pd.DataFrame:
    calc = raw_df.copy()
    calc["month"] = pd.to_datetime(calc["publish_date"], errors="coerce").dt.strftime("%Y-%m").fillna("未知")
    calc["open_rate"] = calc.apply(lambda row: safe_div(row["reads"], row["delivered"]), axis=1)
    calc["share_rate"] = calc.apply(lambda row: safe_div(row["shares"], row["reads"]), axis=1)
    calc["like_rate"] = calc.apply(lambda row: safe_div(row["likes"], row["reads"]), axis=1)
    calc["in_view_rate"] = calc.apply(lambda row: safe_div(row["in_views"], row["reads"]), axis=1)
    calc["favorite_rate"] = calc.apply(lambda row: safe_div(row["favorites"], row["reads"]), axis=1)
    calc["comment_rate"] = calc.apply(lambda row: safe_div(row["comments"], row["reads"]), axis=1)
    calc["follow_rate"] = calc.apply(lambda row: safe_div(row["follows"], row["reads"]), axis=1)
    calc["reward_per_1k_reads"] = calc.apply(lambda row: safe_div(row["reward_yuan"], row["reads"]) * 1000, axis=1)
    calc["goods_per_1k_reads"] = calc.apply(lambda row: safe_div(row["goods_yuan"], row["reads"]) * 1000, axis=1)
    calc["private_reads"] = calc["message_reads"] + calc["chat_reads"] + calc["home_reads"]
    calc["social_reads"] = calc["moments_reads"] + calc["friend_watch_reads"]
    calc["recommend_share"] = calc.apply(lambda row: safe_div(row["recommend_reads"], row["reads"]), axis=1)
    calc["search_share"] = calc.apply(lambda row: safe_div(row["search_reads"], row["reads"]), axis=1)
    calc["private_share"] = calc.apply(lambda row: safe_div(row["private_reads"], row["reads"]), axis=1)
    calc["social_share"] = calc.apply(lambda row: safe_div(row["social_reads"], row["reads"]), axis=1)

    calc["reads_pct"] = percentile_series(calc["reads"])
    calc["open_rate_pct"] = percentile_series(calc["open_rate"])
    calc["share_rate_pct"] = percentile_series(calc["share_rate"])
    calc["like_rate_pct"] = percentile_series(calc["like_rate"])
    calc["in_view_rate_pct"] = percentile_series(calc["in_view_rate"])
    calc["favorite_rate_pct"] = percentile_series(calc["favorite_rate"])
    calc["comment_rate_pct"] = percentile_series(calc["comment_rate"])
    calc["completion_rate_pct"] = percentile_series(calc["completion_rate"])
    calc["avg_stay_seconds_pct"] = percentile_series(calc["avg_stay_seconds"])
    calc["follow_rate_pct"] = percentile_series(calc["follow_rate"])

    calc["reach_score"] = (calc["reads_pct"] * 0.7 + calc["open_rate_pct"] * 0.3).round(1)
    calc["engagement_score"] = (
        calc["share_rate_pct"] * 0.35
        + calc["like_rate_pct"] * 0.25
        + calc["in_view_rate_pct"] * 0.15
        + calc["favorite_rate_pct"] * 0.10
        + calc["comment_rate_pct"] * 0.15
    ).round(1)
    calc["retention_score"] = (calc["completion_rate_pct"] * 0.65 + calc["avg_stay_seconds_pct"] * 0.35).round(1)
    calc["conversion_score"] = calc["follow_rate_pct"].round(1)
    calc["overall_score"] = (
        calc["reach_score"] * 0.4
        + calc["engagement_score"] * 0.3
        + calc["retention_score"] * 0.2
        + calc["conversion_score"] * 0.1
    ).round(1)
    calc["overall_pct"] = percentile_series(calc["overall_score"])
    calc["overall_grade"] = calc["overall_pct"].map(score_to_grade)
    calc["overall_rank"] = calc["overall_score"].rank(method="min", ascending=False).astype(int)
    calc["opportunity_score"] = (calc["retention_score"] - calc["reach_score"]).round(1)
    return calc.sort_values(["overall_score", "reads"], ascending=[False, False]).reset_index(drop=True)


def build_summary_tables(calc_df: pd.DataFrame) -> dict[str, Any]:
    total_reads = float(calc_df["reads"].sum())
    total_articles = int(len(calc_df))
    avg_reads = float(calc_df["reads"].mean()) if total_articles else 0.0
    median_reads = float(calc_df["reads"].median()) if total_articles else 0.0
    avg_score = float(calc_df["overall_score"].mean()) if total_articles else 0.0
    avg_completion = float(calc_df["completion_rate"].mean()) if total_articles else 0.0

    best_article = calc_df.iloc[0].to_dict() if total_articles else {}
    top_articles = calc_df.nlargest(10, "overall_score")[
        ["article_title", "reads", "overall_score", "share_rate", "completion_rate", "follow_rate"]
    ].copy()
    hidden_gems = calc_df.sort_values(["opportunity_score", "retention_score"], ascending=[False, False]).head(10)[
        ["article_title", "reads", "reach_score", "retention_score", "overall_score"]
    ].copy()
    monthly = (
        calc_df.groupby("month", as_index=False)
        .agg(
            article_count=("article_title", "count"),
            total_reads=("reads", "sum"),
            avg_reads=("reads", "mean"),
            avg_score=("overall_score", "mean"),
            avg_completion=("completion_rate", "mean"),
        )
        .sort_values("month")
    )

    channels = pd.DataFrame(
        [
            ("推荐", calc_df["recommend_reads"].sum()),
            ("搜一搜", calc_df["search_reads"].sum()),
            ("公众号消息", calc_df["message_reads"].sum()),
            ("聊天会话", calc_df["chat_reads"].sum()),
            ("公众号主页", calc_df["home_reads"].sum()),
            ("朋友圈", calc_df["moments_reads"].sum()),
            ("朋友在看", calc_df["friend_watch_reads"].sum()),
            ("其他", calc_df["other_reads"].sum()),
        ],
        columns=["channel", "reads"],
    )
    channel_total = channels["reads"].sum() or 1
    channels["share"] = channels["reads"] / channel_total
    channels = channels.sort_values("reads", ascending=False).reset_index(drop=True)

    grades = (
        calc_df["overall_grade"].value_counts()
        .reindex(["S", "A", "B", "C", "D"], fill_value=0)
        .rename_axis("grade")
        .reset_index(name="count")
    )

    return {
        "summary": {
            "total_articles": total_articles,
            "total_reads": total_reads,
            "avg_reads": avg_reads,
            "median_reads": median_reads,
            "avg_score": avg_score,
            "avg_completion": avg_completion,
        },
        "best_article": best_article,
        "top_articles": top_articles,
        "hidden_gems": hidden_gems,
        "monthly": monthly,
        "channels": channels,
        "grades": grades,
    }


def auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 34)


def add_table(ws, display_name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=display_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
        showFirstColumn=False,
        showLastColumn=False,
    )
    ws.add_table(table)


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="123C4A")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(bottom=Side(style="thin", color="D8E1E8"))
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border
    ws.freeze_panes = "A2"


def apply_number_formats(ws, percent_headers: set[str], currency_headers: set[str], date_headers: set[str]) -> None:
    headers = {cell.column: str(cell.value) for cell in ws[1] if cell.value is not None}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            header = headers.get(cell.column)
            if header in percent_headers:
                cell.number_format = "0.00%"
            elif header in currency_headers:
                cell.number_format = '"¥"#,##0.00'
            elif header in date_headers:
                cell.number_format = "yyyy-mm-dd"
            elif header and ("分" in header or "人数" in header or "条数" in header or "阅读" in header or "排名" in header):
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = "0.0" if "分" in header else "#,##0"


def write_raw_sheet(wb: Workbook, raw_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("数据明细")
    headers = [RAW_DISPLAY[column] for column in RAW_COLUMNS]
    ws.append(headers)
    for row in raw_df.itertuples(index=False):
        ws.append(list(row))
    style_header(ws)
    add_table(ws, "RawData")
    apply_number_formats(
        ws,
        percent_headers={"完读率"},
        currency_headers={"赞赏金额(元)", "带货金额(元)"},
        date_headers={"发布日期"},
    )
    auto_fit_columns(ws)


def write_analysis_sheet(wb: Workbook, raw_df: pd.DataFrame) -> None:
    ws = wb.create_sheet("文章分析")
    ws.append(ANALYSIS_HEADERS + HELPER_HEADERS)

    data_row_count = len(raw_df) + 1
    raw_col_lookup = {display: index + 1 for index, display in enumerate(RAW_DISPLAY[column] for column in RAW_COLUMNS)}
    raw_sheet = "数据明细"

    def raw_ref(display_name: str, row_number: int) -> str:
        column_letter = get_column_letter(raw_col_lookup[display_name])
        return f"'{raw_sheet}'!{column_letter}{row_number}"

    for source_row in range(2, data_row_count + 1):
        formula_row = source_row
        formulas = [
            f"={raw_ref('发布日期', source_row)}",
            f'=IF(A{formula_row}="","未知",TEXT(A{formula_row},"yyyy-mm"))',
            f"={raw_ref('文章标题', source_row)}",
            f"={raw_ref('阅读人数', source_row)}",
            f"={raw_ref('送达人数', source_row)}",
            f"=IFERROR(D{formula_row}/E{formula_row},0)",
            f"={raw_ref('完读率', source_row)}",
            f"={raw_ref('平均停留时长(秒)', source_row)}",
            f"={raw_ref('分享人数', source_row)}",
            f"={raw_ref('点赞人数', source_row)}",
            f"={raw_ref('在看人数', source_row)}",
            f"={raw_ref('收藏人数', source_row)}",
            f"={raw_ref('评论条数', source_row)}",
            f"={raw_ref('关注人数', source_row)}",
            f"={raw_ref('赞赏金额(元)', source_row)}",
            f"={raw_ref('带货金额(元)', source_row)}",
            f"=IFERROR(I{formula_row}/D{formula_row},0)",
            f"=IFERROR(J{formula_row}/D{formula_row},0)",
            f"=IFERROR(K{formula_row}/D{formula_row},0)",
            f"=IFERROR(L{formula_row}/D{formula_row},0)",
            f"=IFERROR(M{formula_row}/D{formula_row},0)",
            f"=IFERROR(N{formula_row}/D{formula_row},0)",
            f"=IFERROR(O{formula_row}/D{formula_row}*1000,0)",
            f"=IFERROR(P{formula_row}/D{formula_row}*1000,0)",
            f"=ROUND(AF{formula_row}*0.7+AG{formula_row}*0.3,1)",
            f"=ROUND(AH{formula_row}*0.35+AI{formula_row}*0.25+AJ{formula_row}*0.15+AK{formula_row}*0.10+AL{formula_row}*0.15,1)",
            f"=ROUND(AM{formula_row}*0.65+AN{formula_row}*0.35,1)",
            f"=ROUND(AO{formula_row},1)",
            f"=ROUND(Y{formula_row}*0.4+Z{formula_row}*0.3+AA{formula_row}*0.2+AB{formula_row}*0.1,1)",
            f'=IF(AP{formula_row}>=90,"S",IF(AP{formula_row}>=75,"A",IF(AP{formula_row}>=55,"B",IF(AP{formula_row}>=35,"C","D"))))',
            f"=RANK.EQ(AC{formula_row},$AC$2:$AC${data_row_count},0)",
            f"=IFERROR(PERCENTRANK.INC($D$2:$D${data_row_count},D{formula_row})*100,50)",
            f"=IFERROR(PERCENTRANK.INC($F$2:$F${data_row_count},F{formula_row})*100,50)",
            f"=IFERROR(PERCENTRANK.INC($Q$2:$Q${data_row_count},Q{formula_row})*100,50)",
            f"=IFERROR(PERCENTRANK.INC($R$2:$R${data_row_count},R{formula_row})*100,50)",
            f"=IFERROR(PERCENTRANK.INC($S$2:$S${data_row_count},S{formula_row})*100,50)",
            f"=IFERROR(PERCENTRANK.INC($T$2:$T${data_row_count},T{formula_row})*100,50)",
            f"=IFERROR(PERCENTRANK.INC($U$2:$U${data_row_count},U{formula_row})*100,50)",
            f"=IFERROR(PERCENTRANK.INC($G$2:$G${data_row_count},G{formula_row})*100,50)",
            f"=IFERROR(PERCENTRANK.INC($H$2:$H${data_row_count},H{formula_row})*100,50)",
            f"=IFERROR(PERCENTRANK.INC($V$2:$V${data_row_count},V{formula_row})*100,50)",
            f"=IFERROR(PERCENTRANK.INC($AC$2:$AC${data_row_count},AC{formula_row})*100,50)",
        ]
        ws.append(formulas)

    style_header(ws)
    add_table(ws, "AnalysisData")
    apply_number_formats(
        ws,
        percent_headers={"打开率", "完读率", "分享率", "点赞率", "在看率", "收藏率", "评论率", "关注率"},
        currency_headers={"赞赏金额(元)", "带货金额(元)", "千读赞赏(元)", "千读带货(元)"},
        date_headers={"发布日期"},
    )
    for helper_column in range(len(ANALYSIS_HEADERS) + 1, len(ANALYSIS_HEADERS) + len(HELPER_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(helper_column)].hidden = True

    for score_column in ("Y", "Z", "AA", "AB", "AC"):
        ws.conditional_formatting.add(
            f"{score_column}2:{score_column}{ws.max_row}",
            ColorScaleRule(start_type="num", start_value=0, start_color="F5C0B4", mid_type="num", mid_value=60, mid_color="FCE8A7", end_type="num", end_value=100, end_color="8ED1C1"),
        )

    auto_fit_columns(ws)
    ws.column_dimensions["C"].width = 42


def write_dashboard_sheet(wb: Workbook, calc_df: pd.DataFrame, summary_tables: dict[str, Any]) -> None:
    ws = wb.create_sheet("汇总看板")
    summary = summary_tables["summary"]
    best_article = summary_tables["best_article"]
    top_articles = summary_tables["top_articles"]
    hidden_gems = summary_tables["hidden_gems"]
    monthly = summary_tables["monthly"]
    channels = summary_tables["channels"]
    grades = summary_tables["grades"]

    title_fill = PatternFill("solid", fgColor="123C4A")
    soft_fill = PatternFill("solid", fgColor="F6F2EB")
    key_fill = PatternFill("solid", fgColor="E3F0ED")
    accent_font = Font(color="123C4A", bold=True)

    ws["A1"] = "公众号文章流量分析看板"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:H1")

    ws["A3"] = "统计范围"
    ws["B3"] = f"{summary['total_articles']} 篇文章"
    ws["D3"] = "总阅读"
    ws["E3"] = summary["total_reads"]
    ws["G3"] = "平均综合指数"
    ws["H3"] = summary["avg_score"]

    ws["A4"] = "平均阅读"
    ws["B4"] = summary["avg_reads"]
    ws["D4"] = "阅读中位数"
    ws["E4"] = summary["median_reads"]
    ws["G4"] = "平均完读率"
    ws["H4"] = summary["avg_completion"]

    for cell in ("A3", "D3", "G3", "A4", "D4", "G4"):
        ws[cell].fill = key_fill
        ws[cell].font = accent_font

    ws["A6"] = "本期综合表现最佳"
    ws["A6"].font = Font(bold=True, color="123C4A")
    ws["A7"] = best_article.get("article_title", "-")
    ws["A8"] = "综合指数"
    ws["B8"] = best_article.get("overall_score", 0)
    ws["C8"] = "阅读人数"
    ws["D8"] = best_article.get("reads", 0)
    ws["E8"] = "分享率"
    ws["F8"] = best_article.get("share_rate", 0)
    ws["G8"] = "完读率"
    ws["H8"] = best_article.get("completion_rate", 0)

    start_row = 11
    ws.cell(start_row, 1, "TOP10 文章").font = Font(bold=True, color="123C4A")
    for row in pd.DataFrame(top_articles).itertuples(index=False):
        start_row += 1
        ws.cell(start_row, 1, row.article_title)
        ws.cell(start_row, 5, row.reads)
        ws.cell(start_row, 6, row.overall_score)
        ws.cell(start_row, 7, row.share_rate)
        ws.cell(start_row, 8, row.completion_rate)
    for header, column in zip(["文章标题", "阅读人数", "综合指数", "分享率", "完读率"], [1, 5, 6, 7, 8]):
        ws.cell(12, column, header).fill = soft_fill
        ws.cell(12, column).font = accent_font

    monthly_row = 24
    ws.cell(monthly_row, 1, "月度趋势").font = Font(bold=True, color="123C4A")
    for idx, header in enumerate(["月份", "文章数", "总阅读", "平均阅读", "平均综合指数", "平均完读率"], start=1):
        ws.cell(monthly_row + 1, idx, header).fill = soft_fill
        ws.cell(monthly_row + 1, idx).font = accent_font
    for offset, row in enumerate(monthly.itertuples(index=False), start=monthly_row + 2):
        ws.cell(offset, 1, row.month)
        ws.cell(offset, 2, row.article_count)
        ws.cell(offset, 3, row.total_reads)
        ws.cell(offset, 4, row.avg_reads)
        ws.cell(offset, 5, row.avg_score)
        ws.cell(offset, 6, row.avg_completion)

    channel_row = 24
    channel_col = 8
    ws.cell(channel_row, channel_col, "渠道结构").font = Font(bold=True, color="123C4A")
    for idx, header in enumerate(["渠道", "阅读人数", "占比"], start=channel_col):
        ws.cell(channel_row + 1, idx, header).fill = soft_fill
        ws.cell(channel_row + 1, idx).font = accent_font
    for offset, row in enumerate(channels.itertuples(index=False), start=channel_row + 2):
        ws.cell(offset, channel_col, row.channel)
        ws.cell(offset, channel_col + 1, row.reads)
        ws.cell(offset, channel_col + 2, row.share)

    gem_row = 36
    ws.cell(gem_row, 1, "潜力文章").font = Font(bold=True, color="123C4A")
    for idx, header in enumerate(["文章标题", "阅读人数", "阅读表现分", "留存表现分", "综合指数"], start=1):
        ws.cell(gem_row + 1, idx, header).fill = soft_fill
        ws.cell(gem_row + 1, idx).font = accent_font
    for offset, row in enumerate(hidden_gems.itertuples(index=False), start=gem_row + 2):
        ws.cell(offset, 1, row.article_title)
        ws.cell(offset, 2, row.reads)
        ws.cell(offset, 3, row.reach_score)
        ws.cell(offset, 4, row.retention_score)
        ws.cell(offset, 5, row.overall_score)

    grade_row = 36
    grade_col = 8
    ws.cell(grade_row, grade_col, "等级分布").font = Font(bold=True, color="123C4A")
    for idx, header in enumerate(["等级", "数量"], start=grade_col):
        ws.cell(grade_row + 1, idx, header).fill = soft_fill
        ws.cell(grade_row + 1, idx).font = accent_font
    for offset, row in enumerate(grades.itertuples(index=False), start=grade_row + 2):
        ws.cell(offset, grade_col, row.grade)
        ws.cell(offset, grade_col + 1, row.count)

    apply_number_formats(
        ws,
        percent_headers={"分享率", "完读率", "占比", "平均完读率"},
        currency_headers=set(),
        date_headers=set(),
    )
    auto_fit_columns(ws)
    ws.freeze_panes = "A3"


def write_file_log_sheet(wb: Workbook, file_log: pd.DataFrame) -> None:
    ws = wb.create_sheet("处理日志")
    ws.append(["源文件", "状态", "文章标题", "发布日期", "说明"])
    for row in file_log.itertuples(index=False):
        ws.append(list(row))
    style_header(ws)
    add_table(ws, "FileLog")
    apply_number_formats(ws, percent_headers=set(), currency_headers=set(), date_headers={"发布日期"})
    auto_fit_columns(ws)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["E"].width = 28


def write_formula_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("指标说明")
    ws.append(["指标", "公式", "说明"])
    rows = [
        ("打开率", "阅读人数 / 送达人数", "看推送触达后的打开效率。"),
        ("分享率", "分享人数 / 阅读人数", "看内容被主动扩散的能力。"),
        ("点赞率", "点赞人数 / 阅读人数", "看认可度。"),
        ("在看率", "在看人数 / 阅读人数", "看公开表达意愿。"),
        ("收藏率", "收藏人数 / 阅读人数", "看内容的长期保存价值。"),
        ("评论率", "评论条数 / 阅读人数", "看深度互动意愿。"),
        ("关注率", "关注人数 / 阅读人数", "看文章带来的新增关注转化。"),
        ("千读赞赏(元)", "赞赏金额 / 阅读人数 × 1000", "把赞赏换算到每千阅读，便于横向对比。"),
        ("千读带货(元)", "带货金额 / 阅读人数 × 1000", "把带货金额换算到每千阅读。"),
        ("阅读表现分", "70% 阅读人数分位 + 30% 打开率分位", "兼顾绝对流量和触达效率。"),
        ("互动表现分", "35% 分享率分位 + 25% 点赞率分位 + 15% 在看率分位 + 10% 收藏率分位 + 15% 评论率分位", "更关注文章被传播和被回应的能力。"),
        ("留存表现分", "65% 完读率分位 + 35% 平均停留时长分位", "看文章是否被真正读完、读深。"),
        ("转化表现分", "100% 关注率分位", "统一用关注转化衡量结果型价值。"),
        ("综合指数", "40% 阅读表现分 + 30% 互动表现分 + 20% 留存表现分 + 10% 转化表现分", "用统一口径给每篇文章做总评。"),
        ("综合等级", "按综合指数分位映射 S/A/B/C/D", "S 为前 10%，A 为前 25%，B 为前 45%，C 为前 65%，其余为 D。"),
    ]
    for row in rows:
        ws.append(row)
    style_header(ws)
    add_table(ws, "FormulaGuide")
    auto_fit_columns(ws)
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 34


def export_excel(raw_df: pd.DataFrame, calc_df: pd.DataFrame, file_log: pd.DataFrame, output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    summary_tables = build_summary_tables(calc_df)
    write_dashboard_sheet(wb, calc_df, summary_tables)
    write_raw_sheet(wb, raw_df)
    write_analysis_sheet(wb, raw_df)
    write_formula_sheet(wb)
    write_file_log_sheet(wb, file_log)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_insights(summary_tables: dict[str, Any]) -> list[str]:
    summary = summary_tables["summary"]
    monthly = summary_tables["monthly"]
    channels = summary_tables["channels"]
    best_article = summary_tables["best_article"]
    insights = []

    if not monthly.empty:
        best_month = monthly.sort_values("total_reads", ascending=False).iloc[0]
        insights.append(f"{best_month['month']} 的总阅读最高，累计 {int(best_month['total_reads']):,}。")
    if not channels.empty:
        top_channel = channels.iloc[0]
        insights.append(f"{top_channel['channel']} 是当前最强来源，占全部渠道阅读的 {top_channel['share']:.1%}。")
    if best_article:
        insights.append(
            f"综合指数最高的文章是《{best_article.get('article_title', '-') }》，得分 {best_article.get('overall_score', 0):.1f}。"
        )
    if summary["median_reads"]:
        insights.append(
            f"当前样本的阅读中位数为 {int(summary['median_reads']):,}，比平均阅读更适合做稳定基线。"
        )
    return insights[:4]


def render_bar_rows(items: list[dict[str, Any]], value_key: str, label_key: str, formatter) -> str:
    max_value = max((item[value_key] for item in items), default=1) or 1
    rows = []
    for item in items:
        width = item[value_key] / max_value * 100
        rows.append(
            f"""
            <div class="bar-row">
              <div class="bar-label">{item[label_key]}</div>
              <div class="bar-track"><span style="width:{width:.2f}%"></span></div>
              <div class="bar-value">{formatter(item[value_key])}</div>
            </div>
            """
        )
    return "".join(rows)


def render_table_rows(frame: pd.DataFrame) -> str:
    rows = []
    for row in frame.itertuples(index=False):
        rows.append(
            f"""
            <tr>
              <td>{row.article_title}</td>
              <td>{int(row.reads):,}</td>
              <td>{row.overall_score:.1f}</td>
              <td>{row.share_rate:.2%}</td>
              <td>{row.completion_rate:.2%}</td>
              <td>{row.follow_rate:.2%}</td>
            </tr>
            """
        )
    return "".join(rows)


def export_html(calc_df: pd.DataFrame, output_path: Path) -> None:
    summary_tables = build_summary_tables(calc_df)
    summary = summary_tables["summary"]
    best_article = summary_tables["best_article"]
    top_articles = summary_tables["top_articles"]
    hidden_gems = summary_tables["hidden_gems"]
    monthly_items = summary_tables["monthly"].to_dict("records")
    channel_items = summary_tables["channels"].to_dict("records")
    grade_items = summary_tables["grades"].to_dict("records")
    insights = build_insights(summary_tables)

    monthly_bars = render_bar_rows(monthly_items, "total_reads", "month", lambda value: f"{int(value):,}")
    channel_bars = render_bar_rows(channel_items, "reads", "channel", lambda value: f"{int(value):,}")
    grade_bars = render_bar_rows(grade_items, "count", "grade", lambda value: str(int(value)))
    top_rows = render_table_rows(top_articles)
    gem_rows = "".join(
        f"""
        <tr>
          <td>{row.article_title}</td>
          <td>{int(row.reads):,}</td>
          <td>{row.reach_score:.1f}</td>
          <td>{row.retention_score:.1f}</td>
          <td>{row.overall_score:.1f}</td>
        </tr>
        """
        for row in hidden_gems.itertuples(index=False)
    )

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>公众号文章流量分析报告</title>
  <style>
    :root {{
      --bg: #f3efe8;
      --panel: rgba(255,255,255,0.78);
      --panel-strong: rgba(255,255,255,0.92);
      --ink: #162227;
      --muted: #61727a;
      --line: rgba(22,34,39,0.12);
      --accent: #0f7a66;
      --accent-soft: #d9eee8;
      --warm: #df6d43;
      --shadow: 0 22px 60px rgba(21, 39, 48, 0.10);
      --radius: 22px;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI Variable", "Microsoft YaHei UI", "PingFang SC", sans-serif;
      color: var(--ink);
      background:
        radial-gradient(circle at 10% 20%, rgba(15,122,102,0.16), transparent 28%),
        radial-gradient(circle at 90% 8%, rgba(223,109,67,0.18), transparent 26%),
        linear-gradient(180deg, #f7f2eb 0%, #efe8de 100%);
      min-height: 100vh;
    }}
    .page {{
      width: min(1240px, calc(100% - 32px));
      margin: 24px auto 40px;
    }}
    .hero {{
      background: rgba(255,255,255,0.88);
      color: var(--ink);
      border-radius: 28px;
      padding: 28px;
      box-shadow: var(--shadow);
      position: relative;
      overflow: hidden;
      border: 1px solid rgba(15,122,102,0.12);
    }}
    .hero::after {{
      content: "";
      position: absolute;
      inset: auto -80px -90px auto;
      width: 280px;
      height: 280px;
      border-radius: 50%;
      background: rgba(15,122,102,0.08);
    }}
    .hero h1 {{
      margin: 0 0 10px;
      font-size: clamp(28px, 3vw, 40px);
      line-height: 1.1;
      color: var(--accent);
    }}
    .hero p {{
      margin: 0;
      max-width: 720px;
      color: var(--ink);
      font-size: 15px;
    }}
    .kpis {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 16px;
      margin-top: 18px;
    }}
    .card {{
      background: var(--panel);
      backdrop-filter: blur(10px);
      border: 1px solid rgba(255,255,255,0.38);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
    }}
    .kpi {{
      padding: 18px 18px 16px;
      background: rgba(255, 255, 255, 0.92);
      border: 1px solid rgba(22, 34, 39, 0.08);
      border-radius: 18px;
      color: var(--ink);
      box-shadow: 0 10px 24px rgba(8, 23, 32, 0.08);
    }}
    .kpi .label {{
      color: var(--muted);
      font-size: 13px;
    }}
    .kpi .value {{
      margin-top: 8px;
      font-size: 30px;
      font-weight: 700;
      letter-spacing: -0.03em;
      color: var(--ink);
    }}
    .section-grid {{
      display: grid;
      grid-template-columns: 1.2fr 0.8fr;
      gap: 18px;
      margin-top: 18px;
    }}
    .panel {{
      padding: 20px;
    }}
    .panel h2 {{
      margin: 0 0 14px;
      font-size: 20px;
    }}
    .meta {{
      color: var(--muted);
      font-size: 13px;
    }}
    .insights {{
      margin-top: 18px;
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 14px;
    }}
    .insight {{
      padding: 16px 18px;
      background: var(--panel-strong);
      border-radius: 18px;
      border: 1px solid var(--line);
    }}
    .insight strong {{
      display: block;
      color: var(--accent);
      margin-bottom: 6px;
      font-size: 13px;
      letter-spacing: 0.02em;
      text-transform: uppercase;
    }}
    .feature-title {{
      margin: 6px 0 10px;
      font-size: 24px;
      line-height: 1.35;
    }}
    .badge-row {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      margin-top: 14px;
    }}
    .badge {{
      padding: 8px 12px;
      border-radius: 999px;
      background: var(--accent-soft);
      color: var(--accent);
      font-size: 13px;
      font-weight: 600;
    }}
    .bars {{
      display: grid;
      gap: 12px;
    }}
    .bar-row {{
      display: grid;
      grid-template-columns: 110px minmax(0, 1fr) 88px;
      gap: 12px;
      align-items: center;
    }}
    .bar-label, .bar-value {{
      font-size: 13px;
      color: var(--muted);
    }}
    .bar-track {{
      height: 12px;
      border-radius: 999px;
      background: rgba(15,122,102,0.10);
      overflow: hidden;
      position: relative;
    }}
    .bar-track span {{
      position: absolute;
      inset: 0 auto 0 0;
      border-radius: 999px;
      background: linear-gradient(90deg, var(--accent), #3fa98f);
    }}
    .table-wrap {{
      margin-top: 18px;
      overflow: auto;
      border-radius: 18px;
      border: 1px solid var(--line);
      background: rgba(255,255,255,0.76);
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      min-width: 720px;
    }}
    th, td {{
      padding: 12px 14px;
      border-bottom: 1px solid var(--line);
      text-align: left;
      font-size: 14px;
      vertical-align: top;
    }}
    th {{
      color: var(--muted);
      font-weight: 600;
      background: rgba(18,60,74,0.04);
    }}
    tr:last-child td {{
      border-bottom: none;
    }}
    .search {{
      display: flex;
      justify-content: flex-end;
      margin-top: 14px;
    }}
    .search input {{
      width: min(320px, 100%);
      padding: 11px 14px;
      border-radius: 14px;
      border: 1px solid var(--line);
      background: rgba(255,255,255,0.88);
      font: inherit;
      color: var(--ink);
    }}
    .footer-note {{
      margin-top: 18px;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.7;
    }}
    @media (max-width: 980px) {{
      .kpis, .insights, .section-grid {{
        grid-template-columns: 1fr;
      }}
      .bar-row {{
        grid-template-columns: 90px minmax(0, 1fr) 72px;
      }}
    }}
    @media (prefers-reduced-motion: no-preference) {{
      .card {{
        animation: fadeUp .5s ease both;
      }}
      .card:nth-child(2) {{ animation-delay: .05s; }}
      .card:nth-child(3) {{ animation-delay: .1s; }}
      .card:nth-child(4) {{ animation-delay: .15s; }}
      @keyframes fadeUp {{
        from {{ opacity: 0; transform: translateY(10px); }}
        to {{ opacity: 1; transform: translateY(0); }}
      }}
    }}
  </style>
</head>
<body>
  <div class="page">
    <section class="hero card">
      <h1>公众号文章流量分析报告</h1>
      <p>把“阅读、互动、留存、转化”拆成可理解的四条线，再合成为一张总看板。Excel 负责透明公式，HTML 负责快速读数。</p>
      <div class="kpis">
        <div class="kpi">
          <div class="label">文章数</div>
          <div class="value">{summary["total_articles"]}</div>
        </div>
        <div class="kpi">
          <div class="label">总阅读</div>
          <div class="value">{int(summary["total_reads"]):,}</div>
        </div>
        <div class="kpi">
          <div class="label">平均阅读</div>
          <div class="value">{int(summary["avg_reads"]):,}</div>
        </div>
        <div class="kpi">
          <div class="label">平均综合指数</div>
          <div class="value">{summary["avg_score"]:.1f}</div>
        </div>
      </div>
    </section>

    <div class="insights">
      {''.join(f'<div class="insight card"><strong>Insight</strong><div>{text}</div></div>' for text in insights)}
    </div>

    <div class="section-grid">
      <section class="panel card">
        <div class="meta">综合表现最佳文章</div>
        <div class="feature-title">《{best_article.get("article_title", "-")}》</div>
        <p class="meta">这篇文章在综合指数、阅读效率和互动强度之间取得了最好的平衡。</p>
        <div class="badge-row">
          <span class="badge">综合指数 {best_article.get("overall_score", 0):.1f}</span>
          <span class="badge">阅读 {int(best_article.get("reads", 0)):,}</span>
          <span class="badge">分享率 {best_article.get("share_rate", 0):.2%}</span>
          <span class="badge">完读率 {best_article.get("completion_rate", 0):.2%}</span>
          <span class="badge">关注率 {best_article.get("follow_rate", 0):.2%}</span>
        </div>
      </section>

      <section class="panel card">
        <h2>等级分布</h2>
        <div class="bars">{grade_bars}</div>
      </section>
    </div>

    <div class="section-grid">
      <section class="panel card">
        <h2>月度趋势</h2>
        <div class="bars">{monthly_bars}</div>
      </section>
      <section class="panel card">
        <h2>渠道结构</h2>
        <div class="bars">{channel_bars}</div>
      </section>
    </div>

    <section class="panel card" style="margin-top:18px;">
      <h2>潜力文章</h2>
      <div class="meta">这些文章的留存分高于阅读分，说明内容本身是好的，但传播面还有提升空间。</div>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>文章标题</th>
              <th>阅读人数</th>
              <th>阅读表现分</th>
              <th>留存表现分</th>
              <th>综合指数</th>
            </tr>
          </thead>
          <tbody>{gem_rows}</tbody>
        </table>
      </div>
    </section>

    <section class="panel card" style="margin-top:18px;">
      <h2>TOP10 文章</h2>
      <div class="search"><input id="articleSearch" type="search" placeholder="搜索文章标题"></div>
      <div class="table-wrap">
        <table id="topTable">
          <thead>
            <tr>
              <th>文章标题</th>
              <th>阅读人数</th>
              <th>综合指数</th>
              <th>分享率</th>
              <th>完读率</th>
              <th>关注率</th>
            </tr>
          </thead>
          <tbody>{top_rows}</tbody>
        </table>
      </div>
      <div class="footer-note">
        评分口径：阅读表现分 = 70% 阅读人数分位 + 30% 打开率分位；互动表现分 = 35% 分享率分位 + 25% 点赞率分位 + 15% 在看率分位 + 10% 收藏率分位 + 15% 评论率分位；综合指数 = 40% 阅读表现分 + 30% 互动表现分 + 20% 留存表现分 + 10% 转化表现分。
      </div>
    </section>
  </div>

  <script>
    const searchInput = document.getElementById('articleSearch');
    const rows = Array.from(document.querySelectorAll('#topTable tbody tr'));
    searchInput?.addEventListener('input', (event) => {{
      const keyword = event.target.value.trim().toLowerCase();
      rows.forEach((row) => {{
        row.style.display = row.textContent.toLowerCase().includes(keyword) ? '' : 'none';
      }});
    }});
  </script>
</body>
</html>
"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")


def collect_records(input_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    files = scan_input_files(input_dir)
    signatures: set[str] = set()
    records: list[dict[str, Any]] = []
    file_log: list[dict[str, Any]] = []

    for path in files:
        try:
            file_records, _ = read_file_records(path)
            if not file_records:
                file_log.append(
                    {
                        "源文件": str(path.relative_to(input_dir)),
                        "状态": "跳过",
                        "文章标题": "",
                        "发布日期": None,
                        "说明": "文件中没有识别到文章数据",
                    }
                )
                continue

            for record in file_records:
                record["publish_date"] = coalesce(parse_date(record.get("publish_date")), parse_date_from_path(path))
                record["article_title"] = normalize_text(record.get("article_title")) or path.stem
                record["source_file"] = str(path.relative_to(input_dir))
                signature = build_signature(record)
                if signature in signatures:
                    file_log.append(
                        {
                            "源文件": str(path.relative_to(input_dir)),
                            "状态": "跳过重复",
                            "文章标题": record["article_title"],
                            "发布日期": record["publish_date"],
                            "说明": "与已导入文件内容一致",
                        }
                    )
                    continue
                signatures.add(signature)
                records.append(record)
                file_log.append(
                    {
                        "源文件": str(path.relative_to(input_dir)),
                        "状态": "已导入",
                        "文章标题": record["article_title"],
                        "发布日期": record["publish_date"],
                        "说明": "",
                    }
                )
        except Exception as exc:
            file_log.append(
                {
                    "源文件": str(path.relative_to(input_dir)),
                    "状态": "失败",
                    "文章标题": "",
                    "发布日期": None,
                    "说明": str(exc),
                }
            )

    if not records:
        raise RuntimeError("没有读到可用的公众号文章数据。请检查输入目录中的文件格式。")

    raw_df = coerce_raw_frame(pd.DataFrame(records))
    log_df = pd.DataFrame(file_log, columns=["源文件", "状态", "文章标题", "发布日期", "说明"])
    return raw_df, log_df


def run_analysis(input_dir: Path, output_dir: Path, output_stem: str | None = None) -> AnalysisOutputs:
    raw_df, file_log = collect_records(input_dir)
    calc_df = calculate_metrics(raw_df)

    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = output_stem or f"公众号文章流量分析_{timestamp}"
    excel_path = output_dir / f"{stem}.xlsx"
    html_path = output_dir / f"{stem}.html"

    export_excel(raw_df, calc_df, file_log, excel_path)
    export_html(calc_df, html_path)

    return AnalysisOutputs(
        excel_path=excel_path,
        html_path=html_path,
        total_files=int(len(file_log)),
        imported_articles=int((file_log["状态"] == "已导入").sum()),
        skipped_duplicates=int((file_log["状态"] == "跳过重复").sum()),
        failed_files=int((file_log["状态"] == "失败").sum()),
    )


def main(argv: Iterable[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="整合公众号文章数据并生成 Excel + HTML 分析报告")
    parser.add_argument("--input", "-i", default=str(choose_default_folder("data")), help="输入目录，默认软件所在目录下的 data")
    parser.add_argument("--output-dir", "-o", default=str(choose_default_folder("output")), help="输出目录，默认软件所在目录下的 output")
    parser.add_argument("--name", default="", help="输出文件名前缀，不含扩展名")
    args = parser.parse_args(list(argv) if argv is not None else None)

    input_dir = Path(args.input).expanduser().resolve()
    output_dir = Path(args.output_dir).expanduser().resolve()
    if not input_dir.exists() or not input_dir.is_dir():
        raise SystemExit(f"输入目录不存在: {input_dir}")

    outputs = run_analysis(input_dir, output_dir, args.name or None)
    print(f"Excel: {outputs.excel_path}")
    print(f"HTML:  {outputs.html_path}")
    print(
        f"文件总数 {outputs.total_files}，导入 {outputs.imported_articles}，"
        f"跳过重复 {outputs.skipped_duplicates}，失败 {outputs.failed_files}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
